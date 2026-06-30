import csv
import hmac
import io
import json
import logging
import os
from uuid import uuid4

import httpx
from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    Form,
    Header,
    HTTPException,
    UploadFile,
)
from fastapi.responses import Response
from sqlalchemy import text
from sqlmodel import Session, select

from app.auth import require_auth
from app.config import settings
from app.db import engine
from app.models import EmailResult, Job, User
from app.schemas import BulkJobResponse, BulkStatusResponse
from app.workers.bulk_worker import process_bulk_job

logger = logging.getLogger(__name__)

_VALID_VERDICTS = {"all", "valid", "invalid", "risky", "unknown"}
_MAX_UPLOAD_BYTES = 25 * 1024 * 1024  # 25 MB hard cap on bulk uploads


def _is_privileged(user: User) -> bool:
    return user.role in ("admin", "superadmin")


def _looks_like_xlsx(data: bytes, filename: str) -> bool:
    # XLSX is a ZIP archive — magic bytes "PK\x03\x04"
    return data[:4] == b"PK\x03\x04" or filename.lower().endswith((".xlsx", ".xlsm"))


def _xlsx_to_csv(data: bytes) -> str:
    """Convert the first sheet of an XLSX workbook into CSV text."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(["" if v is None else str(v) for v in row])
    wb.close()
    return buf.getvalue()

router = APIRouter()


def _upload_dir() -> str:
    if settings.upload_dir:
        return settings.upload_dir
    if os.getenv("VERCEL"):
        return "/tmp/uploads"
    return "uploads"


_DISPATCH_HINTS = {
    401: "PAT is rejected (401). Token is invalid or expired — generate a new one.",
    403: "PAT lacks scope (403). Classic PAT needs `workflow` (+ `repo`); fine-grained PAT needs `Actions: Read and write` on this repo.",
    404: "Workflow or repo not found (404). Confirm GITHUB_REPO is `owner/repo` exactly, and that the fine-grained PAT lists this specific repo.",
    422: "Bad inputs (422). Default branch is probably not `main`, or an input is missing — see the body.",
}


async def _count_queued_workflow_runs(
    client: httpx.AsyncClient, workflow_file: str
) -> int:
    """Count GitHub Actions runs of `workflow_file` in `queued` state
    (= dispatched but not yet started — workers are full or the
    concurrency group is waiting). Best-effort: returns 0 on any
    error so a transient GitHub API hiccup never blocks dispatch.
    """
    if not settings.github_pat or not settings.github_repo:
        return 0
    try:
        owner, repo = settings.github_repo.split("/", 1)
    except ValueError:
        return 0
    url = (
        f"https://api.github.com/repos/{owner}/{repo}/"
        f"actions/workflows/{workflow_file}/runs?status=queued&per_page=1"
    )
    try:
        resp = await client.get(
            url,
            headers={
                "Authorization": f"Bearer {settings.github_pat}",
                "Accept": "application/vnd.github+json",
                "X-GitHub-Api-Version": "2022-11-28",
            },
            timeout=4.0,
        )
        if resp.status_code != 200:
            return 0
        return int(resp.json().get("total_count", 0))
    except Exception:  # noqa: BLE001
        return 0


async def _trigger_github_actions(
    job_id: int,
    cache_ttl_days: int | None = None,
    triggered_by: str | None = None,
) -> tuple[bool, str | None]:
    """Dispatch the bulk_process workflow. Returns (ok, error_for_ui)."""
    if not settings.github_pat:
        return False, "GITHUB_PAT env var is not set on Vercel."
    if not settings.github_repo:
        return False, "GITHUB_REPO env var is not set on Vercel."
    try:
        owner, repo = settings.github_repo.split("/", 1)
    except ValueError:
        return False, f"GITHUB_REPO={settings.github_repo!r} is malformed (need owner/repo)."
    url = f"https://api.github.com/repos/{owner}/{repo}/actions/workflows/bulk_process.yml/dispatches"
    inputs: dict[str, str] = {"job_id": str(job_id)}
    if cache_ttl_days is not None:
        inputs["cache_ttl_days"] = str(cache_ttl_days)
    if triggered_by:
        inputs["triggered_by"] = triggered_by
    try:
        async with httpx.AsyncClient(timeout=4.0) as client:
            # Upstream queue gate: refuse if too many runs are already
            # waiting. Workflow concurrency caps in-flight at 3; this
            # caps the wait line so the queue doesn't grow unboundedly.
            cap = settings.max_queued_workflow_runs
            if cap > 0:
                queued = await _count_queued_workflow_runs(client, "bulk_process.yml")
                if queued >= cap:
                    msg = (
                        f"GitHub Actions queue is full: {queued} bulk runs are "
                        f"already waiting (cap: {cap}). Wait for some to start "
                        f"before submitting more."
                    )
                    logger.warning("dispatch refused for job %s: %s", job_id, msg)
                    return False, msg
            resp = await client.post(
                url,
                headers={
                    "Authorization": f"Bearer {settings.github_pat}",
                    "Accept": "application/vnd.github+json",
                    "X-GitHub-Api-Version": "2022-11-28",
                },
                json={"ref": "main", "inputs": inputs},
            )
        if resp.status_code == 204:
            return True, None
        body = resp.text[:200].replace("\n", " ")
        hint = _DISPATCH_HINTS.get(resp.status_code, "See the body and GitHub docs.")
        msg = f"GitHub API returned {resp.status_code}. {hint} Body: {body}"
        logger.warning("dispatch failed for job %s: %s", job_id, msg)
        return False, msg
    except Exception as e:  # noqa: BLE001
        msg = f"GitHub API call raised {type(e).__name__}: {e}"
        logger.warning("dispatch raised for job %s: %s", job_id, msg)
        return False, msg


@router.post("/api/bulk", response_model=BulkJobResponse)
async def create_bulk_job(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    email_column: str = Form(default=""),
    providers: str = Form(default="bouncify"),
    strategy: str = Form(default="bouncify_only"),
    cache_ttl_days: int = Form(default=0),
    current_user: User = Depends(require_auth),
):
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Empty file.")
    if len(contents) > _MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File exceeds {_MAX_UPLOAD_BYTES // (1024 * 1024)} MB limit.",
        )

    # Accept .xlsx / .xlsm by converting to CSV; otherwise decode as text.
    if _looks_like_xlsx(contents, file.filename or ""):
        try:
            csv_str = _xlsx_to_csv(contents)
        except Exception as e:  # noqa: BLE001
            logger.exception("XLSX parse failed: %s", e)
            raise HTTPException(
                status_code=400,
                detail="Could not read the Excel file. Save it as CSV and try again.",
            )
    else:
        try:
            csv_str = contents.decode("utf-8-sig")  # strips BOM if present
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=400,
                detail=(
                    "File is not a readable CSV (looks like binary data). "
                    "Upload a UTF-8 CSV or an .xlsx workbook."
                ),
            )

    # Real row count from the CSV parser (handles missing trailing newline,
    # excludes header, handles embedded newlines inside quoted fields).
    reader = csv.reader(io.StringIO(csv_str))
    parsed_rows = sum(1 for _ in reader)
    row_count = max(0, parsed_rows - 1)
    if settings.max_bulk_emails > 0 and row_count > settings.max_bulk_emails:
        raise HTTPException(
            status_code=400,
            detail=(
                f"File exceeds {settings.max_bulk_emails} email limit per upload."
            ),
        )

    # Per-user concurrency caps: avoid one user saturating the 3-slot
    # GitHub Actions queue and starving everyone else.
    with Session(engine) as session:
        active_jobs, active_emails = session.execute(text(
            "SELECT COUNT(*), COALESCE(SUM(total), 0) FROM job "
            "WHERE user_id = :uid AND status IN ('queued', 'running')"
        ), {"uid": current_user.id}).first() or (0, 0)
    if settings.max_user_active_jobs > 0 and active_jobs >= settings.max_user_active_jobs:
        raise HTTPException(
            status_code=429,
            detail=(
                f"You already have {active_jobs} jobs queued or running "
                f"(limit: {settings.max_user_active_jobs}). Wait for one to finish."
            ),
        )
    if (
        settings.max_user_active_emails > 0
        and active_emails + row_count > settings.max_user_active_emails
    ):
        raise HTTPException(
            status_code=429,
            detail=(
                f"Adding this job would put you over {settings.max_user_active_emails} "
                f"pending emails ({active_emails} already in flight). Wait for one to finish."
            ),
        )

    # Global GHA queue cap: refuse before creating a Job row so the user
    # doesn't end up with a leftover 'failed' job they have to clean up.
    # The same check inside _trigger_github_actions stays as a backstop
    # for the race where the queue fills between this check and dispatch.
    cap = settings.max_queued_workflow_runs
    if cap > 0:
        async with httpx.AsyncClient(timeout=4.0) as client:
            queued = await _count_queued_workflow_runs(client, "bulk_process.yml")
        if queued >= cap:
            raise HTTPException(
                status_code=429,
                detail=(
                    f"GitHub Actions queue is full: {queued} bulk runs are already "
                    f"waiting (cap: {cap}). Wait for some to start before submitting more."
                ),
            )

    upload_dir = _upload_dir()
    os.makedirs(upload_dir, exist_ok=True)
    # uuid4 — collision-free across concurrent uploads; id() reuses across GC.
    filepath = os.path.join(upload_dir, f"upload_{uuid4().hex}.csv")
    with open(filepath, "w", encoding="utf-8", newline="") as f:
        f.write(csv_str)

    with Session(engine) as session:
        job = Job(
            user_id=current_user.id,
            strategy=strategy,
            providers=providers,
            filename=file.filename,
            csv_data=csv_str,
            # Stamp total at upload time so the per-user-emails cap is
            # accurate immediately — the worker re-confirms this number
            # once it parses the CSV itself.
            total=row_count,
        )
        session.add(job)
        session.commit()
        job_id = job.id

    ttl: int | None = cache_ttl_days if cache_ttl_days >= 0 else None

    # IMPORTANT: on Vercel serverless, the function process is terminated as
    # soon as the response is sent — FastAPI BackgroundTasks added at that
    # point do NOT reliably run. So we dispatch INLINE (a fast HTTP POST to
    # GitHub's API, typically <1s) before returning.
    triggered, dispatch_error = await _trigger_github_actions(
        job_id, cache_ttl_days=ttl, triggered_by=current_user.email,
    )
    response_status = "queued"
    if not triggered:
        if os.getenv("VERCEL"):
            reason = dispatch_error or "GitHub Actions dispatch failed (no reason captured)."
            logger.warning("Job %s: %s", job_id, reason)
            with Session(engine) as session:
                job = session.get(Job, job_id)
                if job:
                    job.status = "failed"
                    job.error = reason[:500]
                    session.add(job)
                    session.commit()
            response_status = "failed"
        else:
            background_tasks.add_task(
                process_bulk_job, job_id, filepath, email_column,
                providers.split(","), strategy, ttl,
            )

    return BulkJobResponse(job_id=job_id, total=0, status=response_status)


# Conclusions GitHub Actions reports for a job (see the `if:` context):
#   success, failure, cancelled, skipped, timed_out
_FAILED_CONCLUSIONS = {"failure", "cancelled", "timed_out", "skipped"}


@router.post("/api/bulk/{job_id}/workflow-callback")
async def workflow_callback(
    job_id: int,
    payload: dict,
    x_callback_token: str = Header(default=""),
):
    """Called by the bulk_process workflow's final `if: always()` step so the
    app learns about runs that were cancelled in the GitHub UI, killed by the
    runner host, or timed out — cases where `_mark_failed` inside the script
    never got a chance to run and the Job row would otherwise stay 'running'
    forever.

    Auth is a shared secret in the `X-Callback-Token` header (set
    JOB_CALLBACK_TOKEN on both Vercel and the GitHub repo secrets). No user
    session — the call comes from GitHub Actions, not a browser.
    """
    if not settings.job_callback_token:
        raise HTTPException(status_code=503, detail="JOB_CALLBACK_TOKEN not configured on the server.")
    if not hmac.compare_digest(x_callback_token, settings.job_callback_token):
        raise HTTPException(status_code=401, detail="Bad callback token.")

    conclusion = str(payload.get("conclusion") or "").strip().lower()
    run_url = str(payload.get("run_url") or "").strip()
    reason = str(payload.get("reason") or "").strip()

    with Session(engine) as session:
        job = session.get(Job, job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job not found.")

        # The worker already wrote a final state — don't clobber it.
        if job.status in ("done", "failed"):
            # Still clear csv_data if the worker left it populated.
            if job.csv_data:
                job.csv_data = ""
                session.add(job)
                session.commit()
            return {"ok": True, "noop": True, "status": job.status}

        if conclusion == "success":
            # Worker should have set status=done. If we got here the script
            # exited 0 without committing the final state — surface that.
            job.status = "failed"
            job.error = (f"Workflow finished successfully but the job was never marked done. "
                         f"Run: {run_url}")[:500]
        elif conclusion in _FAILED_CONCLUSIONS:
            human = conclusion.replace("_", " ")
            msg = f"Workflow {human}."
            if reason:
                msg += f" {reason}"
            if run_url:
                msg += f" Run: {run_url}"
            job.status = "failed"
            job.error = msg[:500]
        else:
            # Unknown conclusion — accept the call but record it for triage.
            job.status = "failed"
            job.error = f"Workflow ended with conclusion={conclusion or 'unknown'}. Run: {run_url}"[:500]

        # Drop the raw CSV payload now that the workflow is terminal — the
        # email addresses don't need to live in the DB indefinitely.
        job.csv_data = ""
        session.add(job)
        session.commit()
        return {"ok": True, "status": job.status}


_TEMPLATE_ROWS: list[tuple[str, str, str, str]] = [
    ("email", "name", "source", "notes"),
    ("john.doe@example.com", "John Doe", "website", "replace with your data"),
    ("jane.smith@outlook.com", "Jane Smith", "referral", ""),
    ("user@gmail.com", "Gmail User", "organic", ""),
]


@router.get("/api/bulk/template.xlsx")
def download_xlsx_template():
    """Generate the bulk-upload template as an XLSX on the fly."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "emails"
    for row in _TEMPLATE_ROWS:
        ws.append(row)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="bulk_template.xlsx"'},
    )


@router.get("/api/bulk/template.csv")
def download_csv_template():
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in _TEMPLATE_ROWS:
        writer.writerow(row)
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="bulk_template.csv"'},
    )


@router.get("/api/bulk/{job_id}", response_model=BulkStatusResponse)
async def get_bulk_status(
    job_id: int,
    current_user: User = Depends(require_auth),
):
    with Session(engine) as session:
        job = session.get(Job, job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        if not _is_privileged(current_user) and job.user_id != current_user.id:
            raise HTTPException(status_code=404, detail="Job not found")
        # Aggregate counts at the DB rather than loading every EmailResult row.
        rows = session.execute(
            text("SELECT verdict, COUNT(*) FROM emailresult WHERE job_id = :jid GROUP BY verdict"),
            {"jid": job_id},
        ).fetchall()

    summary: dict[str, int] = {"valid": 0, "invalid": 0, "risky": 0, "unknown": 0}
    for verdict, cnt in rows:
        if verdict in summary:
            summary[verdict] = cnt

    download_url = f"/api/bulk/{job_id}/download" if job.status == "done" else None
    return BulkStatusResponse(
        job_id=job_id,
        status=job.status,
        progress=job.processed,
        total=job.total,
        summary=summary,
        download_url=download_url,
    )


@router.delete("/api/bulk/{job_id}")
async def delete_job(job_id: int, current_user: User = Depends(require_auth)):
    """Delete a single job and all of its EmailResult rows.

    A running job cannot be deleted — its worker would crash mid-write
    against a missing FK row. Mark it failed first if you need to abort.
    """
    if not _is_privileged(current_user):
        raise HTTPException(status_code=403, detail="Admin only")
    with Session(engine) as session:
        job = session.get(Job, job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        if job.status == "running":
            raise HTTPException(
                status_code=409,
                detail="Job is currently running. Wait for it to finish or fail.",
            )
        session.execute(
            text("DELETE FROM emailresult WHERE job_id = :jid"),
            {"jid": job_id},
        )
        session.delete(job)
        session.commit()
    return {"deleted": True, "job_id": job_id}


@router.post("/api/bulk/{job_id}/retry")
async def retry_job(job_id: int, current_user: User = Depends(require_auth)):
    """Re-dispatch a failed job to GitHub Actions.

    Only allowed on jobs in 'failed' status — re-running a queued or running
    job risks two workers writing into the same EmailResult table. We delete
    any partial EmailResult rows from the prior run before queueing the
    fresh dispatch (the worker iterates the full email list every time and
    would otherwise produce duplicate rows).
    """
    with Session(engine) as session:
        job = session.get(Job, job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        if not _is_privileged(current_user) and job.user_id != current_user.id:
            raise HTTPException(status_code=404, detail="Job not found")
        if job.status != "failed":
            raise HTTPException(
                status_code=409,
                detail=f"Only failed jobs can be retried (this one is '{job.status}').",
            )
        if not job.csv_data:
            raise HTTPException(
                status_code=410,
                detail="Original CSV is no longer attached to this job. Re-upload.",
            )
        session.execute(
            text("DELETE FROM emailresult WHERE job_id = :jid"),
            {"jid": job_id},
        )
        job.status = "queued"
        job.processed = 0
        job.error = None
        session.add(job)
        session.commit()

    triggered, dispatch_error = await _trigger_github_actions(
        job_id, triggered_by=current_user.email,
    )
    if not triggered:
        with Session(engine) as session:
            j = session.get(Job, job_id)
            if j:
                j.status = "failed"
                j.error = (dispatch_error or "Retry dispatch failed.")[:500]
                session.add(j)
                session.commit()
        raise HTTPException(status_code=502, detail=dispatch_error or "Dispatch failed.")
    return {"ok": True, "job_id": job_id, "status": "queued"}


@router.post("/api/bulk/clear")
async def clear_all_jobs(current_user: User = Depends(require_auth)):
    """Delete every non-running job (and its EmailResult rows). Admin-only —
    history is shared across users in this app."""
    if not _is_privileged(current_user):
        raise HTTPException(status_code=403, detail="Admin only")
    with Session(engine) as session:
        running = session.execute(
            text("SELECT COUNT(*) FROM job WHERE status = 'running'")
        ).scalar() or 0
        session.execute(text(
            "DELETE FROM emailresult WHERE job_id IN "
            "(SELECT id FROM job WHERE status != 'running')"
        ))
        deleted = session.execute(
            text("DELETE FROM job WHERE status != 'running'")
        ).rowcount or 0
        session.commit()
    return {"deleted": deleted, "kept_running": running}


@router.get("/api/bulk/{job_id}/download")
async def download_bulk(
    job_id: int,
    verdict: str = "all",
    current_user: User = Depends(require_auth),
):
    if verdict not in _VALID_VERDICTS:
        raise HTTPException(status_code=400, detail="Invalid verdict filter.")
    with Session(engine) as session:
        job = session.get(Job, job_id)
        if not job or job.status != "done":
            raise HTTPException(status_code=404, detail="Results not ready")
        if not _is_privileged(current_user) and job.user_id != current_user.id:
            raise HTTPException(status_code=404, detail="Results not ready")
        results = session.exec(
            select(EmailResult).where(EmailResult.job_id == job_id)
        ).all()

    if verdict != "all":
        results = [r for r in results if r.verdict.lower() == verdict.lower()]

    # Build column list from first result's provider_data
    provider_cols: list[str] = []
    if results:
        try:
            pd = json.loads(results[0].provider_data)
            provider_cols = [f"{p}_status" for p in pd]
        except (ValueError, TypeError) as e:
            # Strip CR/LF so a forged provider_data payload can't inject
            # fake log lines (CodeQL py/log-injection). job_id is already
            # int-typed by FastAPI.
            safe_err = str(e).replace("\r", " ").replace("\n", " ")[:200]
            logger.warning(
                "download_bulk: bad provider_data on first row of job %d: %s",
                int(job_id), safe_err,
            )

    fieldnames = ["email", "verdict"] + provider_cols + ["from_cache"]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for r in results:
        row: dict = {"email": r.email, "verdict": r.verdict, "from_cache": False}
        try:
            pd = json.loads(r.provider_data)
            for p, data in pd.items():
                row[f"{p}_status"] = data.get("status", "")
        except (ValueError, TypeError):
            pass
        writer.writerow(row)

    suffix = f"_{verdict}" if verdict != "all" else ""
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="validated_{job_id}{suffix}.csv"'},
    )
